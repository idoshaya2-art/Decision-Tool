from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any, Iterable

from intopia_rules import AREA_CURRENCIES, BASELINE_FX_TO_SF, PLANT_CAPACITY


FINANCE_ALIASES = {
    "revenue_sf": ["revenue", "sales revenue", "net sales", "הכנסות", "מכירות"],
    "gross_profit_sf": ["gross profit", "רווח גולמי"],
    "net_profit_sf": ["net profit", "net income", "רווח נקי"],
    "ending_cash_sf": ["ending cash", "cash balance", "cash", "מזומן סופי", "יתרת מזומן", "מזומן"],
    "debt_sf": ["debt", "loans", "חוב", "הלוואות"],
    "ar_sf": ["accounts receivable", "receivables", "חייבים", "לקוחות"],
    "ap_sf": ["accounts payable", "payables", "זכאים", "ספקים"],
    "research_budget_sf": ["market research", "research budget", "תקציב מחקר", "מחקרי שוק"],
    "rd_x_sf": ["r&d x", "rd x", "מו״פ x", "מופ x"],
    "rd_y_sf": ["r&d y", "rd y", "מו״פ y", "מופ y"],
}

AREA_FINANCE_ALIASES = {
    "area": ["area", "country", "region", "אזור", "מדינה"],
    "currency": ["currency", "מטבע"],
    "fx_to_sf": ["fx", "exchange rate", "fx to sf", "שער חליפין"],
    "revenue_lc": ["revenue", "sales", "הכנסות", "מכירות"],
    "gross_profit_lc": ["gross profit", "רווח גולמי"],
    "net_profit_lc": ["net profit", "net income", "רווח נקי"],
    "ending_cash_lc": ["ending cash", "cash", "מזומן סופי", "מזומן"],
    "debt_lc": ["debt", "loans", "חוב", "הלוואות"],
    "ar_lc": ["accounts receivable", "receivables", "חייבים", "לקוחות"],
    "ap_lc": ["accounts payable", "payables", "זכאים", "ספקים"],
    "inventory_value_lc": ["inventory value", "inventory", "ערך מלאי", "מלאי"],
    "current_assets_lc": ["current assets", "נכסים שוטפים"],
    "current_liabilities_lc": ["current liabilities", "התחייבויות שוטפות"],
    "equity_lc": ["equity", "shareholders equity", "הון עצמי"],
    "total_investment_lc": ["total investment", "invested capital", "השקעה כוללת", "הון מושקע"],
    "operating_cash_flow_lc": ["operating cash flow", "תזרים מפעילות"],
    "capex_commitments_lc": ["capex commitments", "capital commitments", "התחייבויות השקעה"],
}

OPERATION_ALIASES = {
    "area": ["area", "country", "region", "אזור", "מדינה"],
    "product": ["product", "מוצר"],
    "model": ["model", "segment", "דגם", "פלח"],
    "grade": ["grade", "technology", "level", "רמה", "טכנולוגיה"],
    "plants": ["plants", "מפעלים"],
    "plant_capacity": ["capacity", "plant capacity", "קיבולת"],
    "planned_production": ["planned production", "production plan", "ייצור מתוכנן"],
    "actual_production": ["actual production", "production", "ייצור בפועל", "ייצור"],
    "opening_inventory": ["opening inventory", "מלאי פתיחה"],
    "planned_sales": ["planned sales", "sales plan", "מכירות מתוכננות"],
    "actual_sales": ["actual sales", "units sold", "sales units", "מכירות בפועל", "יחידות שנמכרו"],
    "ending_inventory": ["ending inventory", "closing inventory", "מלאי סופי"],
    "forecast_demand": ["forecast demand", "demand forecast", "תחזית ביקוש"],
    "planned_price_lc": ["planned price", "מחיר מתוכנן"],
    "actual_price_lc": ["actual price", "price", "מחיר בפועל", "מחיר"],
    "advertising_lc": ["advertising", "promotion", "פרסום", "קידום"],
    "variable_cost_lc": ["variable cost", "unit cost", "עלות משתנה", "עלות ליחידה"],
    "fixed_cost_lc": ["fixed cost", "עלויות קבועות"],
    "actual_market_share": ["market share", "נתח שוק"],
}


def _clean(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s_\-:/()]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u200f", "").replace("\u200e", "")
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.,\-]", "", text)
    if text.count(",") > 0 and text.count(".") == 0:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else ".".join(parts)
    else:
        text = text.replace(",", "")
    try:
        number = float(text)
        return -abs(number) if negative else number
    except ValueError:
        return None


def _match_field(label: Any, aliases: dict[str, list[str]]) -> str | None:
    normalized = _clean(label)
    if not normalized:
        return None
    for field, options in aliases.items():
        for option in options:
            candidate = _clean(option)
            if normalized == candidate or candidate in normalized:
                return field
    return None


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1255", "windows-1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _csv_rows(content: bytes) -> list[list[Any]]:
    text = _decode(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]


def _xlsx_tables(content: bytes) -> list[tuple[str, list[list[Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel files") from exc
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    tables: list[tuple[str, list[list[Any]]]] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        tables.append((sheet.title, rows))
    return tables


def _pdf_text(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is required to read PDF files") from exc
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _metric_pairs(rows: Iterable[list[Any]]) -> dict[str, float]:
    values: dict[str, float] = {}
    for row in rows:
        cells = [cell for cell in row if cell not in (None, "")]
        if len(cells) < 2:
            continue
        field = _match_field(cells[0], FINANCE_ALIASES)
        value = next((_number(cell) for cell in cells[1:] if _number(cell) is not None), None)
        if field and value is not None:
            values[field] = value
    return values


def _header_map(row: list[Any], aliases: dict[str, list[str]]) -> dict[int, str]:
    result: dict[int, str] = {}
    for index, cell in enumerate(row):
        field = _match_field(cell, aliases)
        if field and field not in result.values():
            result[index] = field
    return result


def _table_records(rows: list[list[Any]], aliases: dict[str, list[str]], required: set[str]) -> list[dict[str, Any]]:
    best_index = -1
    best_map: dict[int, str] = {}
    for index, row in enumerate(rows[:40]):
        mapping = _header_map(row, aliases)
        if len(mapping) > len(best_map):
            best_index, best_map = index, mapping
    if best_index < 0 or not required.issubset(set(best_map.values())):
        return []
    records: list[dict[str, Any]] = []
    text_fields = {"area", "product", "model", "currency"}
    for raw in rows[best_index + 1 :]:
        record: dict[str, Any] = {}
        for column, field in best_map.items():
            if column >= len(raw):
                continue
            value = raw[column]
            if value in (None, ""):
                continue
            record[field] = str(value).strip() if field in text_fields else _number(value)
        if required.issubset({key for key, value in record.items() if value not in (None, "")}):
            records.append(record)
    return records


def _pdf_finance(text: str) -> dict[str, float]:
    result: dict[str, float] = {}
    for line in text.splitlines():
        field = _match_field(line, FINANCE_ALIASES)
        if not field:
            continue
        numbers = re.findall(r"\(?-?[\d][\d,]*(?:\.\d+)?\)?", line)
        if numbers:
            value = _number(numbers[-1])
            if value is not None:
                result[field] = value
    return result


def _strategy_profile_from_text(text: str, filename: str) -> dict[str, Any]:
    lines = [re.sub(r"\s+", " ", line).strip(" \t•-–—") for line in text.splitlines()]
    lines = [line for line in lines if len(line) >= 4]
    if not lines:
        return {}

    def select(*needles: str, limit: int = 12) -> list[str]:
        lowered = [needle.lower() for needle in needles]
        return [line for line in lines if any(needle in line.lower() for needle in lowered)][:limit]

    goals = select("q9", "יעד", "מטרה", "target", "goal", "%", limit=15)
    constraints = select("אין ", "לא ", "מינימום", "מקסימום", "must", "minimum", "maximum", "cash", "מזומן", limit=12)
    priorities = select("מיקוד", "אסטרטג", "עדיפות", "priority", "focus", "טכנולוג", "technology", "שוק", "market", limit=12)
    thesis_lines = [line for line in lines if line not in goals and line not in constraints][:6] or lines[:6]
    return {
        "thesis": " · ".join(thesis_lines)[:2000],
        "priorities": priorities,
        "goals": goals,
        "constraints": constraints,
        "source_name": filename,
        "source_excerpt": "\n".join(lines)[:8000],
    }


def _cell(rows: list[list[Any]], row: int, column: int, default: Any = 0) -> Any:
    """Read a one-based Excel coordinate from a values-only worksheet."""
    try:
        value = rows[row - 1][column - 1]
    except (IndexError, TypeError):
        return default
    return default if value in (None, "") else value


def _sum_cells(rows: list[list[Any]], row_numbers: Iterable[int], columns: Iterable[int]) -> float:
    return sum(_number(_cell(rows, row, column)) or 0 for row in row_numbers for column in columns)


def _intopia_company_number(tables: dict[str, list[list[Any]]]) -> int | None:
    for sheet_name in ("Balance Sheet", "Income Statement", "Management Info"):
        text = str(_cell(tables.get(sheet_name, []), 1, 1, ""))
        match = re.search(r"company\s+(\d+)", text, re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def _detected_quarter(filename: str) -> str:
    match = re.search(r"(?:^|[^A-Z0-9])Q([1-9])(?:[^0-9]|$)", filename.upper())
    return f"Q{match.group(1)}" if match else ""


def _is_intopia_quarter_workbook(tables: list[tuple[str, list[list[Any]]]]) -> bool:
    names = {_clean(name) for name, _ in tables}
    required = {_clean("Balance Sheet"), _clean("Income Statement"), _clean("Management Info"), _clean("Currency")}
    return required.issubset(names)


def _intopia_finance(tables: dict[str, list[list[Any]]]) -> dict[str, Any]:
    balance = tables["Balance Sheet"]
    income = tables["Income Statement"]
    sales_rows = (7, 8, 9, 10, 15, 16, 17, 18)
    interest_bearing_debt = (
        _sum_cells(balance, (29,), (3, 5, 7))
        + _sum_cells(balance, (33, 34), (8,))
    )
    balance_snapshot = {
        "inventory_value_sf": round(_number(_cell(balance, 13, 9)) or 0, 2),
        "current_assets_sf": round(_number(_cell(balance, 16, 9)) or 0, 2),
        "accounts_payable_sf": round(_sum_cells(balance, (26, 27), (9,)), 2),
        "supplier_credit_sf": round(_number(_cell(balance, 28, 9)) or 0, 2),
        "current_liabilities_sf": round(_number(_cell(balance, 30, 9)) or 0, 2),
        "total_assets_sf": round(_number(_cell(balance, 22, 9)) or 0, 2),
        "total_liabilities_sf": round(_number(_cell(balance, 36, 9)) or 0, 2),
        "equity_sf": round(_number(_cell(balance, 47, 9)) or 0, 2),
    }
    return {
        "revenue_sf": round(_sum_cells(income, sales_rows, (9,)), 2),
        "gross_profit_sf": round(_number(_cell(income, 21, 9)) or 0, 2),
        "net_profit_sf": round(_number(_cell(income, 57, 9)) or 0, 2),
        "ending_cash_sf": round(_number(_cell(balance, 6, 9)) or 0, 2),
        # TOTAL LIABILITIES also contains trade payables and supplier credit.
        # Debt is limited to bank loans and long-term bonds.
        "debt_sf": round(interest_bearing_debt, 2),
        "ar_sf": round(_sum_cells(balance, (7, 8), (9,)), 2),
        # Accounts payable is the sum of the two A/P ageing rows. It must
        # never be confused with total current liabilities, which also
        # includes supplier credit and area bank loans.
        "ap_sf": round(_sum_cells(balance, (26, 27), (9,)), 2),
        "research_budget_sf": round(_number(_cell(income, 45, 9)) or 0, 2),
        "rd_x_sf": round(_number(_cell(income, 46, 9)) or 0, 2),
        "rd_y_sf": round(_number(_cell(income, 47, 9)) or 0, 2),
        "dividends_sf": round(_number(_cell(income, 58, 9)) or 0, 2),
        "notes": (
            "חולץ אוטומטית מדוח INTOPIA הרשמי; סכומים מאוחדים ב-SF. "
            f"[[BALANCE:{json.dumps(balance_snapshot, separators=(',', ':'))}]]"
        ),
    }


def _intopia_fx_to_sf(currency_rows: list[list[Any]]) -> dict[str, float]:
    # Each currency block states the current rate into the Liechtenstein/HQ
    # currency in column E. This is the correct conversion into SF.
    result = dict(BASELINE_FX_TO_SF)
    for area, row in {"USA": 14, "Europe": 34, "Brazil": 54, "Liechtenstein": 74}.items():
        value = _number(_cell(currency_rows, row, 5))
        if value and value > 0:
            result[area] = value
    return result


def _intopia_area_finance(tables: dict[str, list[list[Any]]]) -> list[dict[str, Any]]:
    balance = tables["Balance Sheet"]
    income = tables["Income Statement"]
    currency = tables["Currency"]
    fx = _intopia_fx_to_sf(currency)
    sales_rows = (7, 8, 9, 10, 15, 16, 17, 18)
    layout = {
        "USA": {"balance": 3, "income_total": 3, "products": (2, 3)},
        "Europe": {"balance": 5, "income_total": 5, "products": (4, 5)},
        "Brazil": {"balance": 7, "income_total": 7, "products": (6, 7)},
        "Liechtenstein": {"balance": 8, "income_total": 8, "products": ()},
    }
    rows: list[dict[str, Any]] = []
    for area, columns in layout.items():
        balance_column = int(columns["balance"])
        total_column = int(columns["income_total"])
        product_columns = tuple(columns["products"])
        debt_lc = (
            _sum_cells(balance, (33, 34), (balance_column,))
            if area == "Liechtenstein"
            else _sum_cells(balance, (29,), (balance_column,))
        )
        rows.append(
            {
                "area": area,
                "currency": AREA_CURRENCIES[area],
                "fx_to_sf": fx[area],
                "revenue_lc": round(_sum_cells(income, sales_rows, product_columns), 2) if product_columns else 0,
                "gross_profit_lc": round(_sum_cells(income, (21,), product_columns), 2) if product_columns else 0,
                "net_profit_lc": round(_number(_cell(income, 57, total_column)) or 0, 2),
                "ending_cash_lc": round(_number(_cell(balance, 6, balance_column)) or 0, 2),
                "debt_lc": round(debt_lc, 2),
                "ar_lc": round(_sum_cells(balance, (7, 8), (balance_column,)), 2),
                "ap_lc": round(_sum_cells(balance, (26, 27), (balance_column,)), 2),
                "inventory_value_lc": round(_number(_cell(balance, 13, balance_column)) or 0, 2),
                "current_assets_lc": round(_number(_cell(balance, 16, balance_column)) or 0, 2),
                "current_liabilities_lc": round(_number(_cell(balance, 30, balance_column)) or 0, 2),
                "equity_lc": round(_number(_cell(balance, 47, balance_column)) or 0, 2),
                "total_investment_lc": round(_number(_cell(balance, 22, balance_column)) or 0, 2),
                # These fields are not present in the official quarterly
                # workbook. Preserve missingness instead of manufacturing a
                # zero, so the UI and AI cannot interpret "unknown" as none.
                "operating_cash_flow_lc": None,
                "capex_commitments_lc": None,
                "notes": "Actual מתוך Balance Sheet ו-Income Statement.",
            }
        )
    return rows


def _parse_grade_price(value: Any) -> tuple[int | None, float | None]:
    match = re.search(r"[XY]\s*(\d*)\s*:\s*([\d.,]+)", str(value or ""), re.IGNORECASE)
    if not match:
        return None, None
    grade = int(match.group(1)) if match.group(1) else None
    return grade, _number(match.group(2))


def _intopia_market_prices(tables: dict[str, list[list[Any]]], company_number: int | None) -> dict[tuple[str, str, str], dict[str, Any]]:
    rows = tables.get("MR 17&28", [])
    if not rows or company_number is None:
        return {}
    company_row = 0
    for index, row in enumerate(rows, start=1):
        if re.search(rf"company\s+{company_number}\b", str(row[0] if row else ""), re.IGNORECASE):
            company_row = index
            break
    if not company_row:
        return {}
    layout = {
        2: ("USA", "X", "Standard"), 3: ("USA", "X", "Deluxe"),
        4: ("USA", "Y", "Standard"), 5: ("USA", "Y", "Deluxe"),
        6: ("Europe", "X", "Standard"), 7: ("Europe", "X", "Deluxe"),
        8: ("Europe", "Y", "Standard"), 9: ("Europe", "Y", "Deluxe"),
        10: ("Brazil", "X", "Standard"), 11: ("Brazil", "X", "Deluxe"),
        12: ("Brazil", "Y", "Standard"), 13: ("Brazil", "Y", "Deluxe"),
    }
    result: dict[tuple[str, str, str], dict[str, Any]] = {}
    for column, key in layout.items():
        grade, price = _parse_grade_price(_cell(rows, company_row, column, ""))
        if grade is not None or price is not None:
            result[key] = {"grade": grade, "price": price}
    return result


def _intopia_operations(tables: dict[str, list[list[Any]]]) -> list[dict[str, Any]]:
    management = tables["Management Info"]
    income = tables["Income Statement"]
    company_number = _intopia_company_number(tables)
    prices = _intopia_market_prices(tables, company_number)
    fx = _intopia_fx_to_sf(tables["Currency"])
    layout = {
        "USA": {"x": 2, "y": 4, "income_x": 2, "income_y": 3},
        "Europe": {"x": 5, "y": 7, "income_x": 4, "income_y": 5},
        "Brazil": {"x": 8, "y": 10, "income_x": 6, "income_y": 7},
    }
    model_rows = {
        "Standard": {"sales": (7, 8, 9, 10), "cost": (19, 23, 27), "units": (20, 24, 28), "grade": 31, "inventory": 48, "inventory_grade": 49},
        "Deluxe": {"sales": (13, 14, 15, 16), "cost": (21, 25, 29), "units": (22, 26, 30), "grade": 32, "inventory": 50, "inventory_grade": 51},
    }
    result: list[dict[str, Any]] = []
    for area, columns in layout.items():
        office_count = _number(_cell(management, 55, int(columns["x"]))) or 0
        for product, column_key, income_key in (("X", "x", "income_x"), ("Y", "y", "income_y")):
            management_column = int(columns[column_key])
            income_column = int(columns[income_key])
            plants = _number(_cell(management, 53, management_column)) or 0
            capacity = plants * PLANT_CAPACITY[area][product]
            for model, row_map in model_rows.items():
                sales_rows = tuple(row for row in row_map["sales"] if product == "X" or row not in (10, 16))
                actual_sales = _sum_cells(management, sales_rows, (management_column,))
                actual_production = _sum_cells(management, row_map["units"], (management_column,))
                total_cost = _sum_cells(management, row_map["cost"], (management_column,))
                variable_cost = total_cost / actual_production if actual_production else 0
                inventory = _number(_cell(management, int(row_map["inventory"]), management_column)) or 0
                production_grade = _number(_cell(management, int(row_map["grade"]), management_column))
                inventory_grade = _number(_cell(management, int(row_map["inventory_grade"]), management_column))
                market = prices.get((area, product, model), {})
                grade = production_grade if production_grade not in (None, 0) else inventory_grade
                if grade in (None, 0) and market.get("grade") is not None:
                    grade = market["grade"]
                channel_rows = row_map["sales"]
                channel_values = {
                    "consumer": _number(_cell(management, int(channel_rows[0]), management_column)) or 0,
                    "intra_company": _number(_cell(management, int(channel_rows[1]), management_column)) or 0,
                    "inter_company": _number(_cell(management, int(channel_rows[2]), management_column)) or 0,
                    "component": (_number(_cell(management, int(channel_rows[3]), management_column)) or 0) if product == "X" else 0,
                }
                result.append(
                    {
                        "area": area,
                        "fx_to_sf": fx[area],
                        "product": product,
                        "model": model,
                        "grade": int(grade or 0),
                        "plants": plants,
                        "plant_capacity": capacity,
                        "actual_production": actual_production,
                        "actual_sales": actual_sales,
                        "ending_inventory": inventory,
                        "actual_price_lc": market.get("price") or 0,
                        "advertising_lc": (_number(_cell(income, 25, income_column)) or 0) if model == "Standard" else 0,
                        "variable_cost_lc": round(variable_cost, 4),
                        "fixed_cost_lc": (_number(_cell(income, 30, income_column)) or 0) if model == "Standard" else 0,
                        "methods_improvement_lc": (_number(_cell(income, 29, income_column)) or 0) if model == "Standard" else 0,
                        "sales_channel": "משרדי מכירות" if office_count > 0 else "סוכנים",
                        "notes": "Actual channels: " + ", ".join(f"{key}={int(value)}" for key, value in channel_values.items() if value),
                    }
                )
    return result


def _research_area(value: str) -> str:
    normalized = _clean(value)
    if "u.s" in normalized or normalized == "us" or "u s" in normalized:
        return "USA"
    if "ec eu" in normalized or "europe" in normalized:
        return "Europe"
    if "brazil" in normalized:
        return "Brazil"
    return value.strip()


def _intopia_mr17_28(rows: list[list[Any]], quarter: str) -> list[dict[str, Any]]:
    layout = {
        2: ("USA", "X", "Standard"), 3: ("USA", "X", "Deluxe"), 4: ("USA", "Y", "Standard"), 5: ("USA", "Y", "Deluxe"),
        6: ("Europe", "X", "Standard"), 7: ("Europe", "X", "Deluxe"), 8: ("Europe", "Y", "Standard"), 9: ("Europe", "Y", "Deluxe"),
        10: ("Brazil", "X", "Standard"), 11: ("Brazil", "X", "Deluxe"), 12: ("Brazil", "Y", "Standard"), 13: ("Brazil", "Y", "Deluxe"),
    }
    entries: list[dict[str, Any]] = []
    for row_index in range(6, len(rows) + 1):
        company_match = re.search(r"company\s+(\d+)", str(_cell(rows, row_index, 1, "")), re.IGNORECASE)
        if not company_match:
            continue
        for column, (area, product, model) in layout.items():
            grade, price = _parse_grade_price(_cell(rows, row_index, column, ""))
            if grade is None and price is None:
                continue
            entries.append({"company": int(company_match.group(1)), "area": area, "product": product, "model": model, "grade": grade, "price_lc": price})
    grade_entries = [{key: value for key, value in item.items() if key != "price_lc"} for item in entries if item.get("grade") is not None]
    price_entries = [{key: value for key, value in item.items() if key != "grade"} for item in entries if item.get("price_lc") is not None]
    return [
        {"quarter": quarter, "study_id": 17, "title": "MR17 · רמות שנמכרו", "key_result": f"זוהו {len(grade_entries)} הצעות פעילות לפי חברה, אזור, מוצר ודגם.", "numeric_data": {"entries": grade_entries}, "relevance_domains": ["טכנולוגיה", "תמחור", "תחרות"], "confidence": "גבוהה", "status": "מאושר"},
        {"quarter": quarter, "study_id": 28, "title": "MR28 · מחירי מכירה", "key_result": f"זוהו {len(price_entries)} מחירי מתחרים לפי חברה, אזור, מוצר ודגם.", "numeric_data": {"entries": price_entries}, "relevance_domains": ["תמחור", "שיווק", "תחרות"], "confidence": "גבוהה", "status": "מאושר"},
    ]


def _intopia_mr74(rows: list[list[Any]], quarter: str) -> dict[str, Any]:
    companies: list[dict[str, Any]] = []
    for column in range(2, min(12, len(rows[2]) + 1 if len(rows) > 2 else 2)):
        company_match = re.search(r"(\d+)", str(_cell(rows, 3, column, "")))
        if not company_match:
            continue
        companies.append(
            {
                "company": int(company_match.group(1)),
                "cash_k_sf": _number(_cell(rows, 6, column)) or 0,
                "inventory_k_sf": _number(_cell(rows, 9, column)) or 0,
                "assets_k_sf": _number(_cell(rows, 17, column)) or 0,
                "liabilities_k_sf": _number(_cell(rows, 33, column)) or 0,
                "equity_k_sf": _number(_cell(rows, 44, column)) or 0,
                "consumer_sales_k_sf": _number(_cell(rows, 52, column)) or 0,
                "gross_margin_k_sf": _number(_cell(rows, 66, column)) or 0,
                "operating_earnings_k_sf": _number(_cell(rows, 77, column)) or 0,
                "net_earnings_k_sf": _number(_cell(rows, 101, column)) or 0,
            }
        )
    profitable = sum(1 for row in companies if row["net_earnings_k_sf"] > 0)
    return {
        "quarter": quarter,
        "study_id": 74,
        "title": "MR74 · השוואה כספית ענפית",
        "key_result": f"נמצאו נתונים מאוחדים ל-{len(companies)} חברות; {profitable} חברות הציגו רווח נקי חיובי ברבעון.",
        "numeric_data": {"unit": "thousand_SF", "companies": companies},
        "relevance_domains": ["פיננסים", "מגמות", "תחרות", "ציון"],
        "confidence": "גבוהה",
        "status": "מאושר",
    }


def _intopia_mr40(rows: list[list[Any]], quarter: str) -> dict[str, Any]:
    layout = {2: ("USA", "X"), 3: ("USA", "Y"), 4: ("Europe", "X"), 5: ("Europe", "Y"), 6: ("Brazil", "X"), 7: ("Brazil", "Y")}
    entries: list[dict[str, Any]] = []
    for row_index in range(5, len(rows) + 1):
        company_match = re.search(r"company\s+(\d+)", str(_cell(rows, row_index, 1, "")), re.IGNORECASE)
        if not company_match:
            continue
        for column, (area, product) in layout.items():
            entries.append({"company": int(company_match.group(1)), "area": area, "product": product, "plants": _number(_cell(rows, row_index, column)) or 0})
    return {"quarter": quarter, "study_id": 40, "title": "MR40 · מספר מפעלים", "key_result": f"מופו מפעלי {len({row['company'] for row in entries})} חברות לפי אזור ומוצר.", "numeric_data": {"entries": entries}, "relevance_domains": ["קיבולת", "ייצור", "תחרות"], "confidence": "גבוהה", "status": "מאושר"}


def _intopia_range_research(rows: list[list[Any]], quarter: str, study_id: int) -> dict[str, Any]:
    ranges: list[dict[str, Any]] = []
    for row in rows:
        text = " ".join(str(cell) for cell in row if cell not in (None, ""))
        match = re.search(r"from\s+([\d.]+)\s+to\s+([\d.]+).*area\s+(.+?)\s+relative", text, re.IGNORECASE)
        if match:
            ranges.append({"area": _research_area(match.group(3)), "low_pct": float(match.group(1)), "high_pct": float(match.group(2))})
    product_match = re.search(r"(?:chip|pc)\s+grade\s*\((\d+)\)", " ".join(str(cell) for row in rows[:3] for cell in row if cell), re.IGNORECASE)
    product = "Y" if any("PC GRADE" in str(cell).upper() for row in rows[:3] for cell in row if cell) else "X"
    grade = int(product_match.group(1)) if product_match else max(0, study_id - (60 if product == "Y" else 50))
    is_price = 31 <= study_id <= 49
    kind = "תוספת מחיר" if is_price else "תוספת עלות ייצור"
    summary = ", ".join(f"{row['area']} {row['low_pct']:.0f}–{row['high_pct']:.0f}%" for row in ranges)
    return {
        "quarter": quarter,
        "study_id": study_id,
        "title": f"MR{study_id} · {kind} {product}{grade}",
        "area": "",
        "product": product,
        "key_result": summary or "לא זוהה טווח מספרי; יש לבדוק את תצוגת המקור.",
        "numeric_data": {"product": product, "grade": grade, "measure": "price_premium_pct" if is_price else "cost_premium_pct", "ranges": ranges},
        "relevance_domains": ["תמחור", "Unit Economics", "טכנולוגיה"],
        "confidence": "גבוהה" if ranges else "בינונית",
        "status": "מאושר",
    }


def _intopia_research(tables: dict[str, list[list[Any]]], quarter: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    if tables.get("MR 17&28"):
        result.extend(_intopia_mr17_28(tables["MR 17&28"], quarter))
    if tables.get("MR 74"):
        result.append(_intopia_mr74(tables["MR 74"], quarter))
    for sheet_name, rows in tables.items():
        compact = re.sub(r"\s+", "", sheet_name).upper()
        match = re.fullmatch(r"MR(\d+)", compact)
        if not match or not rows:
            continue
        study_id = int(match.group(1))
        if study_id in {17, 28, 74}:
            continue
        meaningful = [cell for row in rows for cell in row if cell not in (None, "")]
        if len(meaningful) < 2:
            continue
        if study_id == 40:
            result.append(_intopia_mr40(rows, quarter))
        elif 31 <= study_id <= 69:
            result.append(_intopia_range_research(rows, quarter, study_id))
        else:
            result.append(
                {
                    "quarter": quarter,
                    "study_id": study_id,
                    "title": f"MR{study_id}",
                    "key_result": " · ".join(str(value) for value in meaningful[:12])[:1800],
                    "numeric_data": {"rows": [row[:20] for row in rows[:80]]},
                    "relevance_domains": ["מחקר שוק"],
                    "confidence": "בינונית",
                    "status": "מאושר",
                }
            )
    return result


def _intopia_validation(tables: dict[str, list[list[Any]]], operations: list[dict[str, Any]]) -> list[dict[str, str]]:
    balance = tables["Balance Sheet"]
    checks: list[dict[str, str]] = []
    assets = _number(_cell(balance, 22, 9)) or 0
    liabilities_equity = _number(_cell(balance, 49, 9)) or 0
    difference = round(assets - liabilities_equity, 2)
    checks.append({"code": "balance_reconciliation", "status": "ok" if abs(difference) <= 1 else "error", "message": f"מאזן מאוחד: הפרש {difference:,.2f} SF."})
    for row in operations:
        production = _number(row.get("actual_production")) or 0
        capacity = _number(row.get("plant_capacity")) or 0
        if production > capacity + 1 and capacity >= 0:
            checks.append({"code": "capacity", "status": "warning", "message": f"{row['area']} {row['product']} {row['model']}: ייצור {production:,.0f} מעל קיבולת {capacity:,.0f}."})
    if not any(check["status"] in {"error", "warning"} for check in checks):
        checks.append({"code": "structure", "status": "ok", "message": "מבנה הדוח וכל בדיקות העקביות המרכזיות תקינים."})
    return checks


def _extract_intopia_workbook(tables_list: list[tuple[str, list[list[Any]]]], filename: str, quarter: str) -> tuple[dict[str, Any], list[str]]:
    tables = {name: rows for name, rows in tables_list}
    operations = _intopia_operations(tables)
    detected = _detected_quarter(filename)
    validation = _intopia_validation(tables, operations)
    issues = [check["message"] for check in validation if check["status"] == "error"]
    if detected and quarter in {f"Q{i}" for i in range(1, 10)} and detected != quarter:
        issues.append(f"שם הקובץ מצביע על {detected}, אך הוא שויך ל-{quarter}.")
    extracted = {
        "finance": _intopia_finance(tables),
        "finance_by_area": _intopia_area_finance(tables),
        "operations": operations,
        "research_results": _intopia_research(tables, quarter),
        "strategy_profile": {},
        "previews": [{"sheet": name, "rows": rows[:8]} for name, rows in tables_list],
        "metadata": {"detected_quarter": detected or quarter, "company_number": _intopia_company_number(tables), "source_format": "INTOPIA quarterly output"},
        "validation": validation,
    }
    return extracted, issues


def extract_document(content: bytes, filename: str, mime_type: str, quarter: str, category: str) -> dict[str, Any]:
    extension = Path(filename).suffix.lower()
    extracted: dict[str, Any] = {"finance": {}, "finance_by_area": [], "operations": [], "research_results": [], "strategy_profile": {}, "previews": []}
    issues: list[str] = []
    parser = "unsupported"
    category_key = category.lower()
    is_strategy = "אסטרטג" in category_key or "strategy" in category_key or "יעדי q9" in category_key or "q9 goals" in category_key

    try:
        if extension in {".xlsx", ".xlsm"} or "spreadsheet" in mime_type:
            parser = "excel"
            tables = _xlsx_tables(content)
            if not is_strategy and _is_intopia_quarter_workbook(tables):
                parser = "intopia-quarterly-v1"
                extracted, exact_issues = _extract_intopia_workbook(tables, filename, quarter)
                issues.extend(exact_issues)
            else:
                strategy_text: list[str] = []
                for sheet_name, rows in tables:
                    extracted["finance"].update(_metric_pairs(rows))
                    extracted["finance_by_area"].extend(_table_records(rows, AREA_FINANCE_ALIASES, {"area"}))
                    extracted["operations"].extend(_table_records(rows, OPERATION_ALIASES, {"area", "product", "model"}))
                    extracted["previews"].append({"sheet": sheet_name, "rows": rows[:8]})
                    if is_strategy:
                        strategy_text.extend(str(cell) for row in rows[:250] for cell in row if cell not in (None, ""))
                if is_strategy:
                    extracted["strategy_profile"] = _strategy_profile_from_text("\n".join(strategy_text), filename)
        elif extension in {".csv", ".tsv"} or mime_type.startswith("text/csv"):
            parser = "csv"
            rows = _csv_rows(content)
            extracted["finance"].update(_metric_pairs(rows))
            extracted["finance_by_area"].extend(_table_records(rows, AREA_FINANCE_ALIASES, {"area"}))
            extracted["operations"].extend(_table_records(rows, OPERATION_ALIASES, {"area", "product", "model"}))
            extracted["previews"].append({"sheet": "CSV", "rows": rows[:12]})
        elif extension == ".pdf" or mime_type == "application/pdf":
            parser = "pdf-text"
            text = _pdf_text(content)
            if not text.strip():
                issues.append("לא נמצא טקסט קריא. ייתכן שזה PDF סרוק ונדרש פענוח AI/OCR.")
            extracted["finance"].update(_pdf_finance(text))
            extracted["previews"].append({"sheet": "PDF", "text": text[:4000]})
            if is_strategy:
                extracted["strategy_profile"] = _strategy_profile_from_text(text, filename)
            if "מחקר" in category or "research" in category.lower():
                extracted["research_results"].append({"quarter": quarter, "title": filename, "key_result": text[:2000], "confidence": "בינונית" if text.strip() else "נמוכה"})
        elif mime_type.startswith("text/") or extension in {".txt", ".md"}:
            parser = "text"
            text = _decode(content)
            extracted["finance"].update(_pdf_finance(text))
            extracted["previews"].append({"sheet": "Text", "text": text[:4000]})
            if is_strategy:
                extracted["strategy_profile"] = _strategy_profile_from_text(text, filename)
            if "מחקר" in category or "research" in category.lower():
                extracted["research_results"].append({"quarter": quarter, "title": filename, "key_result": text[:2000], "confidence": "בינונית"})
        elif mime_type.startswith("image/"):
            parser = "image"
            issues.append("התמונה נשמרה, אך נדרש פענוח AI/OCR כדי לחלץ ממנה נתונים.")
        else:
            issues.append("סוג הקובץ נשמר אך אינו נתמך עדיין בחילוץ אוטומטי.")
    except Exception as exc:
        issues.append(f"החילוץ נכשל: {exc}")

    detected_quarter = str((extracted.get("metadata") or {}).get("detected_quarter") or "")
    resolved_quarter = quarter if quarter in {f"Q{i}" for i in range(1, 10)} else detected_quarter
    if resolved_quarter not in {f"Q{i}" for i in range(1, 10)} and (extracted["finance"] or extracted["operations"]):
        issues.append("נמצאו נתונים רבעוניים אך הקובץ לא קושר לרבעון Q1–Q9.")
    if not extracted["finance"] and not extracted["operations"] and not extracted["research_results"] and not extracted["strategy_profile"]:
        issues.append("לא זוהו שדות מובנים. הקובץ נשמר ויופיע לבדיקה ידנית.")

    extracted_count = len(extracted["finance"]) + len(extracted["finance_by_area"]) + len(extracted["operations"]) + len(extracted["research_results"]) + (1 if extracted["strategy_profile"] else 0)
    confidence = "גבוהה" if extracted_count >= 8 and not issues else "בינונית" if extracted_count >= 2 else "נמוכה"
    status = "מוכן לאישור" if extracted_count else "נדרשת בדיקה"
    return {"parser_type": parser, "status": status, "confidence": confidence, "extracted_data": extracted, "issues": issues, "extracted_count": extracted_count}
