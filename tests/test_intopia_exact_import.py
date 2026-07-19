from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from import_service import extract_document


def _quarter_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title in ("Balance Sheet", "Income Statement", "Management Info", "Currency", "MR 74", "MR 17&28", "MR41", "MR61"):
        workbook.create_sheet(title)

    balance = workbook["Balance Sheet"]
    balance["A1"] = "COMPANY 8"
    balance["I6"] = 1_250_000
    balance["I22"] = 4_000_000
    balance["C29"] = 200_000
    balance["E29"] = 300_000
    balance["G29"] = 400_000
    balance["H33"] = 500_000
    balance["H34"] = 100_000
    balance["I33"] = 500_000
    balance["I34"] = 100_000
    balance["I36"] = 1_700_000
    balance["I49"] = 4_000_000
    income = workbook["Income Statement"]
    income["I10"] = 600_000
    income["I21"] = 180_000
    income["I45"] = 35_000
    income["I46"] = 500_000
    income["I47"] = 300_000
    income["I57"] = -100_000
    currency = workbook["Currency"]
    for row, rate in ((14, 1.1), (34, 1.5), (54, 0.48), (74, 1)):
        currency.cell(row=row, column=5, value=rate)
    management = workbook["Management Info"]
    management["E53"] = 1
    management["E20"] = 35_000
    management["E48"] = 35_000
    management["E49"] = 0

    mr_prices = workbook["MR 17&28"]
    mr_prices["A6"] = "COMPANY 8"
    mr_prices["F6"] = "X0:40"
    mr_prices["H6"] = "Y1:165"
    mr74 = workbook["MR 74"]
    mr74["B3"] = "COMPANY 8"
    mr74["B101"] = -100
    for sheet, low_high in (("MR41", ((26, 35), (20, 28), (17, 26))), ("MR61", ((9, 16), (7, 12), (8, 13)) )):
        ws = workbook[sheet]
        ws["A1"] = f"MARKET RESEARCH NUMBER {sheet[2:]}"
        ws["A2"] = "MARKETING TEST FOR PC GRADE (1)" if sheet == "MR41" else "ESTIMATE OF VARIABLE MANUFACTURING COST FOR PC GRADE (1)"
        for index, (area, values) in enumerate(zip(("U.S.", "EC/EU", "BRAZIL"), low_high), start=3):
            ws.cell(row=index, column=1, value=f"ESTIMATED CHANGE WILL GO FROM {values[0]}. TO {values[1]}. PERCENT IN AREA {area} RELATIVE TO GRADE ZERO.")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_exact_intopia_parser_detects_quarter_and_paid_research():
    result = extract_document(
        _quarter_workbook(),
        "Q2 Real Simulation.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Setup",
        "פלט רבעוני",
    )
    extracted = result["extracted_data"]
    assert result["parser_type"] == "intopia-quarterly-v1"
    assert result["confidence"] == "גבוהה"
    assert extracted["metadata"]["detected_quarter"] == "Q2"
    assert extracted["metadata"]["company_number"] == 8
    assert extracted["finance"]["revenue_sf"] == 600_000
    assert extracted["finance"]["debt_sf"] == 1_500_000
    assert '"total_assets_sf":4000000.0' in extracted["finance"]["notes"]
    assert len(extracted["finance_by_area"]) == 4
    debts = {row["area"]: row["debt_lc"] for row in extracted["finance_by_area"]}
    assert debts == {"USA": 200_000, "Europe": 300_000, "Brazil": 400_000, "Liechtenstein": 600_000}
    assert len(extracted["operations"]) == 12
    studies = {row["study_id"]: row for row in extracted["research_results"]}
    assert studies[41]["numeric_data"]["ranges"][1] == {"area": "Europe", "low_pct": 20.0, "high_pct": 28.0}
    assert studies[61]["numeric_data"]["ranges"][2] == {"area": "Brazil", "low_pct": 8.0, "high_pct": 13.0}


def test_complete_decision_catalog_is_exposed(client):
    response = client.get("/api/scenario-actions")
    assert response.status_code == 200
    actions = response.json()
    assert len(actions) == 19
    assert {row["code"] for row in actions} >= {"A1-1", "A2-4", "A4", "H1-2", "H5", "H6", "W1", "W2", "W3"}
